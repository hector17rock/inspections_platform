import openpyxl
import os

path = r"C:\Users\HRS0044.s03716\Documents\Availability  WK 21.xlsx"
print("File exists:", os.path.exists(path))
wb = openpyxl.load_workbook(path)
print("Sheet names:", wb.sheetnames)
sheet = wb.active
print("Active sheet title:", sheet.title)
for r in range(1, 50):
    row_vals = [sheet.cell(r, c).value for c in range(1, 12)]
    if any(x is not None for x in row_vals):
        # Truncate values to avoid too much noise
        truncated = [str(x)[:40] if x is not None else "" for x in row_vals]
        print(f"Row {r:02d}: {truncated}")
